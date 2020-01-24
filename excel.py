from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.styles import Font

'''print(get_column_letter(2))
print(column_index_from_string('D'))'''

twb=Workbook()
#twb=load_workbook('t.xlsx')



tws=twb.active
tws['A1']=2
tws['E5']=42
tws.append([1,2,3,4])
tws.title='tws'
print(type(tws['A1']))
'''print(tuple(tws.rows))
for row in tws.iter_cols(1,4,2,3):
    for cell in row:
        print(cell)
print(tws.max_row)
tws.row_dimensions[7].height=40'''
tws.merge_cells('A1:E5')
tws.unmerge_cells('A1:E5')

tws1=twb.create_sheet('newSheet1')
tws2=twb.create_sheet('newSheet2')
tws3=twb.create_sheet('newSheet3')
tws4=twb.create_sheet('newSheet4')

tws5=twb.copy_worksheet('Formula')

twb.save('t.xlsx')
